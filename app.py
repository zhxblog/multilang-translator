import os
import json
import asyncio
import io
import base64
from typing import Optional

import anthropic
import pandas as pd
import openpyxl
from openai import OpenAI
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="多语言翻译插件")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

LANGUAGE_MAP = {
    "en":    "English",
    "de":    "German",
    "fr":    "French",
    "ru":    "Russian",
    "pt":    "Portuguese",
    "ko":    "Korean",
    "ja":    "Japanese",
    "es":    "Spanish",
    "ar":    "Arabic",
    "it":    "Italian",
    "bn":    "Bengali",
    "da":    "Danish",
    "idn":   "Indonesian",
    "th":    "Thai",
    "vi":    "Vietnamese",
    "tr":    "Turkish",
    "nl":    "Dutch",
    "pl":    "Polish",
    "zh_HN": "Traditional Chinese",
    "w":     "Czech",
    "ee":    "Estonian",
    "fi":    "Finnish",
    "el":    "Greek",
    "hu":    "Hungarian",
    "gg":    "Bulgarian",
    "hr":    "Croatian",
    "ga":    "Irish",
    "ro":    "Romanian",
    "lv":    "Latvian",
    "mt":    "Maltese",
    "sk":    "Slovak",
    "si":    "Slovenian",
    "sv":    "Swedish",
    "lt":    "Lithuanian",
    "ml":    "Malay",
    "bur":   "Burmese (Myanmar)",
    "kh":    "Khmer (Cambodian)",
    "la":    "Lao",
    "ph":    "Filipino (Tagalog)",
    "fa":    "Persian (Farsi)",
    "uk":    "Ukrainian",
    "ta":    "Tamil",
    "mn":    "Mongolian",
    "kk":    "Kazakh",
}

LANG_NAMES_ZH = {
    "en":    "英语",
    "de":    "德语",
    "fr":    "法语",
    "ru":    "俄语",
    "pt":    "葡萄牙语",
    "ko":    "韩语",
    "ja":    "日语",
    "es":    "西班牙语",
    "ar":    "阿拉伯语",
    "it":    "意大利语",
    "bn":    "孟加拉语",
    "da":    "丹麦语",
    "idn":   "印尼语",
    "th":    "泰语",
    "vi":    "越南语",
    "tr":    "土耳其语",
    "nl":    "荷兰语",
    "pl":    "波兰语",
    "zh_HN": "中文繁体",
    "w":     "捷克语",
    "ee":    "爱沙尼亚语",
    "fi":    "芬兰语",
    "el":    "希腊语",
    "hu":    "匈牙利语",
    "gg":    "保加利亚语",
    "hr":    "克罗地亚语",
    "ga":    "爱尔兰语",
    "ro":    "罗马尼亚语",
    "lv":    "拉脱维亚语",
    "mt":    "马耳他语",
    "sk":    "斯洛伐克语",
    "si":    "斯洛文尼亚语",
    "sv":    "瑞典语",
    "lt":    "立陶宛语",
    "ml":    "马来语",
    "bur":   "缅甸语",
    "kh":    "柬埔寨语",
    "la":    "老挝语",
    "ph":    "菲律宾语（他加禄语）",
    "fa":    "波斯语",
    "uk":    "乌克兰语",
    "ta":    "泰米尔语",
    "mn":    "蒙古语",
    "kk":    "哈萨克语",
}

TRANSLATE_PROMPT = """Translate the following Simplified Chinese text into multiple languages.
Return ONLY a valid JSON object where keys are the language codes and values are the translations.
Do not include any explanation, markdown, or extra text - just the raw JSON.

Chinese text: "{text}"

Target languages (code: language name):
{lang_list}

Example output format:
{{"en": "translation here", "de": "Übersetzung hier", ...}}"""


def _extract_json(text: str) -> dict:
    text = text.strip()
    start = text.find("{")
    end = text.rfind("}") + 1
    if start >= 0 and end > start:
        text = text[start:end]
    return json.loads(text)


def translate_anthropic(api_key: str, model: str, text: str, target_langs: list[str]) -> dict:
    lang_list = "\n".join([f'  "{c}": "{LANGUAGE_MAP[c]}"' for c in target_langs if c in LANGUAGE_MAP])
    prompt = TRANSLATE_PROMPT.format(text=text, lang_list=lang_list)
    client = anthropic.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model=model,
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    return _extract_json(msg.content[0].text)


PROVIDER_BASE_URLS = {
    "aliyun_coding": "https://coding.dashscope.aliyuncs.com/v1",
    "aliyun":        "https://dashscope.aliyuncs.com/compatible-mode/v1",
}


def translate_openai_compat(base_url: str, api_key: str, model: str, text: str, target_langs: list[str]) -> dict:
    lang_list = "\n".join([f'  "{c}": "{LANGUAGE_MAP[c]}"' for c in target_langs if c in LANGUAGE_MAP])
    prompt = TRANSLATE_PROMPT.format(text=text, lang_list=lang_list)
    import httpx
    client = OpenAI(
        api_key=api_key,
        base_url=base_url,
        timeout=httpx.Timeout(120.0, connect=15.0),
    )
    resp = client.chat.completions.create(
        model=model,
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    return _extract_json(resp.choices[0].message.content)


def translate_one(provider: str, api_key: str, model: str, text: str, target_langs: list[str]) -> dict:
    if provider == "anthropic":
        return translate_anthropic(api_key, model, text, target_langs)
    base_url = PROVIDER_BASE_URLS.get(provider, PROVIDER_BASE_URLS["aliyun_coding"])
    return translate_openai_compat(base_url, api_key, model, text, target_langs)


def build_excel(original_bytes: bytes, translations: dict[str, dict[str, str]]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    ws = wb.active

    col_to_idx = {}
    zh_col = None

    # Excel row 1 has codes (zh_CN, en...), row 2 has Chinese names (中文简体, 英语...)
    for col_idx in range(1, ws.max_column + 1):
        r1 = str(ws.cell(row=1, column=col_idx).value or "").strip()
        r2 = str(ws.cell(row=2, column=col_idx).value or "").strip()
        if r1 in ("中文简体", "zh_CN") or r2 == "中文简体":
            zh_col = col_idx
        for code in LANGUAGE_MAP:
            if r1 == code or r2 == LANG_NAMES_ZH.get(code, "") or r1 == LANG_NAMES_ZH.get(code, ""):
                col_to_idx[code] = col_idx
                break

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx in range(3, ws.max_row + 1):
        if zh_col is None:
            continue
        zh_val = ws.cell(row=row_idx, column=zh_col).value
        if zh_val and str(zh_val).strip() in translations:
            trans = translations[str(zh_val).strip()]
            for code, col in col_to_idx.items():
                if code in trans and trans[code]:
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = trans[code]
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 40)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@app.get("/", response_class=HTMLResponse)
async def index():
    return HTML_CONTENT


@app.post("/translate-stream")
async def translate_stream(request: Request):
    form = await request.form()
    api_key  = form.get("api_key", "").strip()
    provider = form.get("provider", "aliyun_coding").strip()
    model    = form.get("model", "").strip()
    file     = form.get("file")

    if not api_key:
        raise HTTPException(400, "请提供 API Key")
    if not file:
        raise HTTPException(400, "请上传 Excel 文件")

    # Default models
    defaults = {"anthropic": "claude-sonnet-4-6", "aliyun_coding": "qwen3.5-plus", "aliyun": "qwen-plus"}
    if not model:
        model = defaults.get(provider, "qwen3.5-plus")

    file_bytes = await file.read()

    async def gen():
        try:
            yield f"data: {json.dumps({'type': 'status', 'message': '正在解析 Excel 文件...'})}\n\n"

            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None)
            zh_col_idx = 0
            for i in range(len(df.columns)):
                v = str(df.iloc[0, i]) if pd.notna(df.iloc[0, i]) else ""
                if v in ("中文简体", "zh_CN"):
                    zh_col_idx = i
                    break

            # Row 0: column codes (zh_CN, en, ...)
            # Row 1: column labels (中文简体, 英语, ...) — skip this header label row
            # Row 2+: actual content to translate
            zh_texts = [
                str(df.iloc[r, zh_col_idx]).strip()
                for r in range(2, len(df))
                if pd.notna(df.iloc[r, zh_col_idx]) and str(df.iloc[r, zh_col_idx]).strip()
            ]

            # Check both row 0 (codes like "en") and row 1 (Chinese names like "英语")
            existing_codes = []
            for col_i in range(len(df.columns)):
                h0 = str(df.iloc[0, col_i]) if pd.notna(df.iloc[0, col_i]) else ""
                h1 = str(df.iloc[1, col_i]) if len(df) > 1 and pd.notna(df.iloc[1, col_i]) else ""
                for code in LANGUAGE_MAP:
                    if h0 == code or h1 == LANG_NAMES_ZH.get(code, "") or h0 == LANG_NAMES_ZH.get(code, ""):
                        existing_codes.append(code)
                        break
            target_langs = [c for c in existing_codes if c in LANGUAGE_MAP]

            total = len(zh_texts)
            provider_label = {"aliyun_coding": "阿里云 Coding Plan", "aliyun": "阿里云 DashScope"}.get(provider, "Anthropic Claude")
            yield f"data: {json.dumps({'type': 'status', 'message': f'使用 {provider_label} ({model})，共 {total} 条文本，翻译为 {len(target_langs)} 种语言...'})}\n\n"

            all_translations = {}
            loop = asyncio.get_event_loop()
            success_count = 0

            for i, text in enumerate(zh_texts):
                yield f"data: {json.dumps({'type': 'progress', 'current': i+1, 'total': total, 'text': text})}\n\n"
                try:
                    trans = await loop.run_in_executor(
                        None,
                        lambda t=text: translate_one(provider, api_key, model, t, target_langs)
                    )
                    all_translations[text] = trans
                    success_count += 1
                    yield f"data: {json.dumps({'type': 'translated', 'text': text, 'langs': len(trans)})}\n\n"
                except Exception as e:
                    err_msg = str(e)
                    # 鉴权失败立即中止，不继续浪费请求
                    is_auth_error = any(kw in err_msg for kw in (
                        "401", "invalid_api_key", "Incorrect API key",
                        "AuthenticationError", "authentication_error",
                    ))
                    if is_auth_error:
                        if provider == "aliyun_coding":
                            tip = "Coding Plan API Key 不正确，请在阿里云 Coding Plan 控制台「套餐专属 API Key」处重新复制"
                        elif provider == "aliyun":
                            tip = "阿里云 DashScope API Key 不正确，请到 DashScope 控制台重新获取"
                        else:
                            tip = "Anthropic API Key 不正确，请到 https://console.anthropic.com/ 获取正确的 Key"
                        yield f"data: {json.dumps({'type': 'fatal', 'message': f'❌ API Key 验证失败：{tip}'})}\n\n"
                        return
                    yield f"data: {json.dumps({'type': 'error', 'text': text, 'message': err_msg})}\n\n"
                    all_translations[text] = {}

            if success_count == 0:
                yield f"data: {json.dumps({'type': 'fatal', 'message': '所有翻译均失败，未生成文件。请检查 API Key 和网络连接。'})}\n\n"
                return

            yield f"data: {json.dumps({'type': 'status', 'message': f'翻译完成（成功 {success_count}/{total} 条），正在生成 Excel...'})}\n\n"
            output_bytes = build_excel(file_bytes, all_translations)
            encoded = base64.b64encode(output_bytes).decode()
            yield f"data: {json.dumps({'type': 'done', 'file': encoded, 'filename': '翻译结果.xlsx'})}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'type': 'fatal', 'message': str(e)})}\n\n"

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


HTML_CONTENT = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>多语言翻译插件</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --primary: #4f46e5;
    --primary-hover: #4338ca;
    --success: #10b981;
    --error: #ef4444;
    --bg: #f8fafc;
    --surface: #ffffff;
    --border: #e2e8f0;
    --text: #1e293b;
    --muted: #64748b;
    --radius: 12px;
  }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif;
    background: var(--bg); color: var(--text);
    min-height: 100vh; padding: 2rem 1rem;
  }
  .container { max-width: 760px; margin: 0 auto; }
  .header { text-align: center; margin-bottom: 2.5rem; }
  .header h1 {
    font-size: 2rem; font-weight: 700;
    background: linear-gradient(135deg, #4f46e5, #7c3aed);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    margin-bottom: .5rem;
  }
  .header p { color: var(--muted); font-size: .95rem; }
  .card {
    background: var(--surface); border-radius: var(--radius);
    border: 1px solid var(--border); padding: 1.75rem;
    margin-bottom: 1.5rem; box-shadow: 0 1px 3px rgba(0,0,0,.06);
  }
  .card-title {
    font-size: .85rem; font-weight: 600; text-transform: uppercase;
    letter-spacing: .05em; color: var(--muted); margin-bottom: 1rem;
  }
  .form-group { margin-bottom: 1.25rem; }
  label { display: block; font-size: .875rem; font-weight: 500; margin-bottom: .4rem; }

  /* Provider tabs */
  .tabs { display: flex; gap: .5rem; margin-bottom: 1.25rem; }
  .tab {
    flex: 1; padding: .6rem 1rem; border-radius: 8px; border: 2px solid var(--border);
    background: var(--surface); cursor: pointer; font-size: .875rem; font-weight: 500;
    color: var(--muted); transition: all .15s; text-align: center;
  }
  .tab:hover { border-color: var(--primary); color: var(--primary); }
  .tab.active { border-color: var(--primary); background: rgba(79,70,229,.07); color: var(--primary); }
  .tab .tab-icon { font-size: 1.2rem; display: block; margin-bottom: .2rem; }

  input[type="text"], input[type="password"], select {
    width: 100%; padding: .625rem .875rem; border: 1px solid var(--border);
    border-radius: 8px; font-size: .9rem; outline: none;
    transition: border-color .15s, box-shadow .15s;
    background: var(--surface); color: var(--text);
  }
  input:focus, select:focus {
    border-color: var(--primary);
    box-shadow: 0 0 0 3px rgba(79,70,229,.12);
  }
  .api-key-wrap { position: relative; }
  .toggle-key {
    position: absolute; right: .75rem; top: 50%; transform: translateY(-50%);
    background: none; border: none; cursor: pointer; color: var(--muted); font-size: .8rem;
  }
  .hint { font-size: .78rem; color: var(--muted); margin-top: .35rem; }
  a.link { color: var(--primary); text-decoration: none; }
  a.link:hover { text-decoration: underline; }

  .drop-zone {
    border: 2px dashed var(--border); border-radius: 10px;
    padding: 2rem; text-align: center; cursor: pointer;
    transition: border-color .2s, background .2s; position: relative;
  }
  .drop-zone:hover, .drop-zone.drag-over {
    border-color: var(--primary); background: rgba(79,70,229,.04);
  }
  .drop-zone input[type="file"] { position: absolute; inset: 0; opacity: 0; cursor: pointer; }
  .drop-icon { font-size: 2.5rem; margin-bottom: .75rem; }
  .drop-zone p { color: var(--muted); font-size: .9rem; }
  .drop-zone p span { color: var(--primary); font-weight: 500; }
  .file-info {
    display: none; align-items: center; gap: .75rem;
    padding: .75rem 1rem; background: rgba(79,70,229,.06);
    border-radius: 8px; margin-top: .75rem; font-size: .875rem;
  }
  .file-info.show { display: flex; }
  .file-name { flex: 1; font-weight: 500; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .file-size { color: var(--muted); font-size: .8rem; }
  .file-rm { background: none; border: none; cursor: pointer; color: var(--error); font-size: 1rem; }

  .langs-grid {
    display: grid; grid-template-columns: repeat(auto-fill, minmax(130px, 1fr));
    gap: .4rem; max-height: 200px; overflow-y: auto; padding: .25rem;
  }
  .lang-tag {
    padding: .3rem .6rem; background: rgba(79,70,229,.08);
    border-radius: 6px; font-size: .78rem; color: var(--primary); font-weight: 500;
  }

  .btn {
    display: inline-flex; align-items: center; justify-content: center;
    gap: .5rem; padding: .7rem 1.5rem; border: none; border-radius: 8px;
    font-size: .95rem; font-weight: 600; cursor: pointer; transition: all .15s; width: 100%;
  }
  .btn-primary { background: var(--primary); color: white; }
  .btn-primary:hover:not(:disabled) { background: var(--primary-hover); }
  .btn-primary:disabled { opacity: .55; cursor: not-allowed; }
  .btn-success { background: var(--success); color: white; margin-top: 1rem; }
  .btn-success:hover { background: #059669; }

  .progress-wrap { display: none; }
  .progress-wrap.show { display: block; }
  .progress-bar-bg {
    height: 8px; background: var(--border); border-radius: 99px;
    overflow: hidden; margin: .75rem 0;
  }
  .progress-bar {
    height: 100%; background: linear-gradient(90deg, var(--primary), #7c3aed);
    border-radius: 99px; transition: width .3s ease; width: 0%;
  }
  .progress-text { font-size: .85rem; color: var(--muted); display: flex; justify-content: space-between; }
  .log-box {
    margin-top: 1rem; max-height: 220px; overflow-y: auto;
    background: #0f172a; border-radius: 8px; padding: .875rem 1rem;
    font-family: 'SF Mono', 'Fira Code', monospace; font-size: .8rem; line-height: 1.6;
  }
  .log-line { padding: .1rem 0; }
  .log-line.ok { color: #86efac; }
  .log-line.info { color: #93c5fd; }
  .log-line.err { color: #fca5a5; }

  .result-card { display: none; text-align: center; padding: 2rem; }
  .result-card.show { display: block; }
  .result-icon { font-size: 3rem; margin-bottom: 1rem; }
  .result-card h3 { font-size: 1.2rem; margin-bottom: .5rem; }
  .result-card p { color: var(--muted); font-size: .875rem; }

  @media (max-width: 480px) {
    .header h1 { font-size: 1.5rem; }
    .card { padding: 1.25rem; }
  }
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>🌐 多语言翻译插件</h1>
    <p>上传 Excel 文件，一键翻译为 44 种语言</p>
  </div>

  <!-- Provider & API -->
  <div class="card">
    <div class="card-title">API 配置</div>

    <div class="form-group">
      <label>选择 AI 提供商</label>
      <div class="tabs">
        <div class="tab active" id="tab-aliyun_coding" onclick="switchProvider('aliyun_coding')">
          <span class="tab-icon">⚡</span>阿里云 Coding Plan
        </div>
        <div class="tab" id="tab-aliyun" onclick="switchProvider('aliyun')">
          <span class="tab-icon">☁️</span>阿里云 DashScope
        </div>
        <div class="tab" id="tab-anthropic" onclick="switchProvider('anthropic')">
          <span class="tab-icon">🤖</span>Anthropic Claude
        </div>
      </div>
    </div>

    <div class="form-group" id="model-group">
      <label for="modelSelect" id="model-label">选择模型</label>
      <select id="modelSelect"></select>
      <p class="hint" id="model-hint"></p>
    </div>

    <div class="form-group">
      <label for="apiKey" id="key-label">API Key</label>
      <div class="api-key-wrap">
        <input type="password" id="apiKey" placeholder="" autocomplete="off">
        <button class="toggle-key" onclick="toggleKey()">👁</button>
      </div>
      <p class="hint" id="key-hint"></p>
    </div>
  </div>

  <!-- File Upload -->
  <div class="card">
    <div class="card-title">上传文件</div>
    <div class="drop-zone" id="dropZone"
         ondragover="onDragOver(event)" ondragleave="onDragLeave()" ondrop="onDrop(event)">
      <input type="file" id="fileInput" accept=".xlsx,.xls" onchange="onFileSelect(this)">
      <div class="drop-icon">📊</div>
      <p>拖放 Excel 文件到这里，或 <span>点击选择</span></p>
      <p style="margin-top:.35rem;font-size:.8rem">支持 .xlsx / .xls</p>
    </div>
    <div class="file-info" id="fileInfo">
      <span>📄</span>
      <span class="file-name" id="fileName"></span>
      <span class="file-size" id="fileSize"></span>
      <button class="file-rm" onclick="removeFile()">✕</button>
    </div>
  </div>

  <!-- Languages -->
  <div class="card">
    <div class="card-title">目标语言（44 种）</div>
    <div class="langs-grid" id="langsGrid"></div>
  </div>

  <!-- Start -->
  <div class="card">
    <button class="btn btn-primary" id="translateBtn" onclick="startTranslation()">
      ✨ 开始翻译
    </button>
  </div>

  <!-- Progress -->
  <div class="card progress-wrap" id="progressWrap">
    <div class="card-title">翻译进度</div>
    <div class="progress-bar-bg"><div class="progress-bar" id="progressBar"></div></div>
    <div class="progress-text">
      <span id="progressLabel">准备中...</span>
      <span id="progressCount">0 / 0</span>
    </div>
    <div class="log-box" id="logBox"></div>
  </div>

  <!-- Result -->
  <div class="card result-card" id="resultCard">
    <div class="result-icon">🎉</div>
    <h3>翻译完成！</h3>
    <p id="resultSummary"></p>
    <button class="btn btn-success" onclick="downloadResult()">⬇️ 下载翻译结果 (.xlsx)</button>
  </div>
</div>

<script>
const PROVIDERS = {
  aliyun_coding: {
    label: '套餐专属 API Key',
    placeholder: 'sk-sp-...',
    hint: '在阿里云 Coding Plan 控制台「套餐专属 API Key」处复制，Base URL 已自动设置为 <code>https://coding.dashscope.aliyuncs.com/v1</code>',
    models: [
      { value: 'qwen3.5-plus',        label: 'qwen3.5-plus（推荐，支持视觉）' },
      { value: 'qwen3-max-2026-01-23', label: 'qwen3-max-2026-01-23（最强）' },
      { value: 'qwen3-coder-plus',     label: 'qwen3-coder-plus（编程优化）' },
      { value: 'qwen3-coder-next',     label: 'qwen3-coder-next（编程优化）' },
      { value: 'glm-5',               label: 'glm-5（智谱，深度思考）' },
      { value: 'glm-4.7',             label: 'glm-4.7（智谱）' },
      { value: 'kimi-k2.5',           label: 'kimi-k2.5（Kimi，视觉理解）' },
      { value: 'MiniMax-M2.5',        label: 'MiniMax-M2.5' },
    ],
    modelHint: '翻译任务推荐 qwen3.5-plus，速度快质量好；qwen3-max 效果最强',
  },
  aliyun: {
    label: 'API Key（标准 DashScope）',
    placeholder: 'sk-xxxxxxxxxxxx',
    hint: '在阿里云 DashScope 控制台获取标准 API Key',
    models: [
      { value: 'qwen-plus',  label: 'qwen-plus（推荐）' },
      { value: 'qwen-turbo', label: 'qwen-turbo（最快）' },
      { value: 'qwen-max',   label: 'qwen-max（最强）' },
    ],
    modelHint: '标准 DashScope API，Base URL: dashscope.aliyuncs.com',
  },
  anthropic: {
    label: 'Anthropic API Key',
    placeholder: 'sk-ant-...',
    hint: '在 Anthropic 控制台获取：<a class="link" href="https://console.anthropic.com/" target="_blank">console.anthropic.com</a>',
    models: [
      { value: 'claude-opus-4-6',          label: 'claude-opus-4-6（最强）' },
      { value: 'claude-sonnet-4-6',        label: 'claude-sonnet-4-6（推荐）' },
      { value: 'claude-haiku-4-5-20251001', label: 'claude-haiku-4-5（最快）' },
    ],
    modelHint: 'Sonnet 翻译质量与速度平衡最佳',
  },
};

const LANGUAGES = [
  ["en","英语"],["de","德语"],["fr","法语"],["ru","俄语"],["pt","葡萄牙语"],
  ["ko","韩语"],["ja","日语"],["es","西班牙语"],["ar","阿拉伯语"],["it","意大利语"],
  ["bn","孟加拉语"],["da","丹麦语"],["idn","印尼语"],["th","泰语"],["vi","越南语"],
  ["tr","土耳其语"],["nl","荷兰语"],["pl","波兰语"],["zh_HN","中文繁体"],["w","捷克语"],
  ["ee","爱沙尼亚语"],["fi","芬兰语"],["el","希腊语"],["hu","匈牙利语"],["gg","保加利亚语"],
  ["hr","克罗地亚语"],["ga","爱尔兰语"],["ro","罗马尼亚语"],["lv","拉脱维亚语"],["mt","马耳他语"],
  ["sk","斯洛伐克语"],["si","斯洛文尼亚语"],["sv","瑞典语"],["lt","立陶宛语"],["ml","马来语"],
  ["bur","缅甸语"],["kh","柬埔寨语"],["la","老挝语"],["ph","菲律宾语"],["fa","波斯语"],
  ["uk","乌克兰语"],["ta","泰米尔语"],["mn","蒙古语"],["kk","哈萨克语"]
];

let currentProvider = 'aliyun';
let selectedFile = null;
let resultBlob = null;

// Init language tags
const grid = document.getElementById('langsGrid');
LANGUAGES.forEach(([, name]) => {
  const t = document.createElement('div');
  t.className = 'lang-tag';
  t.textContent = name;
  grid.appendChild(t);
});

function switchProvider(p) {
  currentProvider = p;
  ['aliyun_coding', 'aliyun', 'anthropic'].forEach(id => {
    const el = document.getElementById('tab-' + id);
    if (el) el.classList.toggle('active', p === id);
  });

  const cfg = PROVIDERS[p];
  document.getElementById('key-label').textContent = cfg.label;
  document.getElementById('apiKey').placeholder = cfg.placeholder;
  document.getElementById('key-hint').innerHTML = cfg.hint;
  document.getElementById('model-hint').textContent = cfg.modelHint;

  const sel = document.getElementById('modelSelect');
  sel.innerHTML = cfg.models.map(m => `<option value="${m.value}">${m.label}</option>`).join('');
}

// Init
switchProvider('aliyun_coding');

function toggleKey() {
  const i = document.getElementById('apiKey');
  i.type = i.type === 'password' ? 'text' : 'password';
}

function onDragOver(e) { e.preventDefault(); document.getElementById('dropZone').classList.add('drag-over'); }
function onDragLeave() { document.getElementById('dropZone').classList.remove('drag-over'); }
function onDrop(e) { e.preventDefault(); onDragLeave(); if (e.dataTransfer.files[0]) setFile(e.dataTransfer.files[0]); }
function onFileSelect(i) { if (i.files[0]) setFile(i.files[0]); }
function setFile(f) {
  selectedFile = f;
  document.getElementById('fileName').textContent = f.name;
  document.getElementById('fileSize').textContent = fmt(f.size);
  document.getElementById('fileInfo').classList.add('show');
}
function removeFile() {
  selectedFile = null;
  document.getElementById('fileInput').value = '';
  document.getElementById('fileInfo').classList.remove('show');
}
function fmt(b) {
  if (b < 1024) return b + ' B';
  if (b < 1048576) return (b/1024).toFixed(1) + ' KB';
  return (b/1048576).toFixed(1) + ' MB';
}

function log(msg, type='info') {
  const box = document.getElementById('logBox');
  const d = document.createElement('div');
  d.className = 'log-line ' + type;
  d.textContent = '[' + new Date().toLocaleTimeString() + '] ' + msg;
  box.appendChild(d);
  box.scrollTop = box.scrollHeight;
}

function setProgress(cur, tot) {
  const pct = tot > 0 ? Math.round(cur/tot*100) : 0;
  document.getElementById('progressBar').style.width = pct + '%';
  document.getElementById('progressCount').textContent = cur + ' / ' + tot;
}

async function startTranslation() {
  const apiKey = document.getElementById('apiKey').value.trim();
  const model  = document.getElementById('modelSelect').value;
  if (!apiKey) { alert('请输入 API Key'); return; }
  if (!selectedFile) { alert('请选择 Excel 文件'); return; }

  const btn = document.getElementById('translateBtn');
  btn.disabled = true; btn.textContent = '翻译中...';
  document.getElementById('progressWrap').classList.add('show');
  document.getElementById('resultCard').classList.remove('show');
  document.getElementById('logBox').innerHTML = '';
  resultBlob = null;

  const fd = new FormData();
  fd.append('api_key', apiKey);
  fd.append('provider', currentProvider);
  fd.append('model', model);
  fd.append('file', selectedFile);

  try {
    const resp = await fetch('/translate-stream', { method: 'POST', body: fd });
    if (!resp.ok) {
      const e = await resp.json().catch(() => ({ detail: resp.statusText }));
      throw new Error(e.detail || '请求失败');
    }
    const reader = resp.body.getReader();
    const dec = new TextDecoder();
    let buf = '';
    while (true) {
      const { done, value } = await reader.read();
      if (done) break;
      buf += dec.decode(value, { stream: true });
      const lines = buf.split('\\n');
      buf = lines.pop();
      for (const line of lines) {
        if (!line.startsWith('data: ')) continue;
        try { handle(JSON.parse(line.slice(6))); } catch {}
      }
    }
  } catch (e) {
    log('错误: ' + e.message, 'err');
    alert('翻译失败: ' + e.message);
  }
  btn.disabled = false; btn.textContent = '✨ 开始翻译';
}

function handle(evt) {
  switch (evt.type) {
    case 'status':
      log(evt.message, 'info');
      document.getElementById('progressLabel').textContent = evt.message;
      break;
    case 'progress':
      setProgress(evt.current, evt.total);
      document.getElementById('progressLabel').textContent = '翻译中: ' + evt.text;
      log('[' + evt.current + '/' + evt.total + '] ' + evt.text, 'info');
      break;
    case 'translated':
      log('✓ ' + evt.text + ' (' + evt.langs + ' 种语言)', 'ok');
      break;
    case 'error':
      log('✗ ' + evt.text + ': ' + evt.message, 'err');
      break;
    case 'fatal':
      log('严重错误: ' + evt.message, 'err');
      document.getElementById('progressLabel').textContent = '❌ 翻译失败';
      document.getElementById('translateBtn').disabled = false;
      document.getElementById('translateBtn').textContent = '✨ 开始翻译';
      break;
    case 'done':
      log('✅ 全部完成，正在准备文件...', 'ok');
      setProgress(100, 100);
      document.getElementById('progressLabel').textContent = '完成！';
      const bin = atob(evt.file);
      const bytes = new Uint8Array(bin.length);
      for (let i = 0; i < bin.length; i++) bytes[i] = bin.charCodeAt(i);
      resultBlob = new Blob([bytes], {
        type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
      });
      document.getElementById('resultSummary').textContent = '文件已准备好，点击下方按钮下载';
      document.getElementById('resultCard').classList.add('show');
      break;
  }
}

function downloadResult() {
  if (!resultBlob) return;
  const url = URL.createObjectURL(resultBlob);
  const a = document.createElement('a');
  a.href = url; a.download = '翻译结果.xlsx'; a.click();
  URL.revokeObjectURL(url);
}
</script>
</body>
</html>"""


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("app:app", host="0.0.0.0", port=port, reload=False)
